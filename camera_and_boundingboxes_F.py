from ultralytics import YOLO
import cv2
import time
import threading

import numpy as np
import subprocess
from working_cam_sensor.vl53l5cx_sensor import VL53L5CXSensor

# Optional Excel logging support; fall back gracefully if openpyxl is not installed
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    _has_openpyxl = True
except ImportError:
    Workbook = None  # type: ignore
    Font = None  # type: ignore
    PatternFill = None  # type: ignore
    Alignment = None  # type: ignore
    _has_openpyxl = False

from datetime import datetime


def _sanitize_speech_text(text: str) -> str:
    """Sanitize text for PowerShell speech synthesis (escape single quotes)."""
    return text.replace("'", "''")


# Voice settings (Windows TTS)
VOICE_NAME = "Microsoft Zira Desktop"  # female voice on Windows
TTS_PROCESS = None


def speak_text(text: str) -> None:
    """Speak text using Windows PowerShell TTS (System.Speech).

    This function ensures only one speech process is active at a time.
    """
    global TTS_PROCESS

    # Do not start a new speech process if one is still running.
    if TTS_PROCESS is not None and TTS_PROCESS.poll() is None:
        return

    try:
        safe_text = _sanitize_speech_text(text)
        cmd = [
            "powershell",
            "-Command",
            (
                "Add-Type -AssemblyName System.Speech; "
                "$s = New-Object System.Speech.Synthesis.SpeechSynthesizer; "
                f"$s.SelectVoice('{VOICE_NAME}'); "
                "$s.Rate = 2; "          # -10 (slowest) to 10 (fastest); 4 = noticeably faster
                f"$s.Speak('{safe_text}');"
            )
        ]
        # Fire and forget -- do not block main loop
        TTS_PROCESS = subprocess.Popen(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception as e:
        # Fallback: print error once, do not keep spamming
        if not hasattr(speak_text, "_warned"):
            print(f"ALERT: TTS failed ({e})")
            speak_text._warned = True


def get_direction_descriptor(cx: int, cy: int, frame_w: int, frame_h: int) -> str:
    """Return a human readable direction (e.g., upper right, center) for a point in the frame."""
    # Divide the frame into a 3x3 grid
    third_w = frame_w / 3.0
    third_h = frame_h / 3.0

    if cx < third_w:
        horiz = "left"
    elif cx < 2 * third_w:
        horiz = "center"
    else:
        horiz = "right"

    if cy < third_h:
        vert = "upper"
    elif cy < 2 * third_h:
        vert = "center"
    else:
        vert = "bottom"

    if vert == "center" and horiz == "center":
        return "center"
    if vert == "center":
        return horiz
    if horiz == "center":
        return vert
    return f"{vert} {horiz}"


def _euclidean_dist(p0, p1):
    return ((p0[0] - p1[0]) ** 2 + (p0[1] - p1[1]) ** 2) ** 0.5


def _find_tracked_object(label: str, center: tuple[int, int]) -> dict | None:
    """Find a previously tracked object with the same label and nearby location."""
    for obj in tracked_objects:
        if obj["label"] != label:
            continue
        if _euclidean_dist(center, obj["center"]) < TRACKED_OBJECT_MOVE_THRESHOLD_PX:
            return obj
    return None


def _cleanup_tracked_objects(now: float) -> None:
    """Remove tracked objects that have not been seen recently."""
    global tracked_objects
    tracked_objects = [obj for obj in tracked_objects if (now - obj.get("last_seen", 0)) < TRACKED_OBJECT_MAX_AGE]


# Create YOLO model (suppress verbose output to avoid flooding the terminal)
model = YOLO("yolov10n.pt", verbose=False)

cap = cv2.VideoCapture(0)
if not cap.isOpened():
    print("ALERT: camera could not be opened")
    cap = None

# Initialize VL53L5CX sensor
# For serial communication (ESP32), use: sensor = VL53L5CXSensor(port='COM3')  # Replace with your COM port
# For direct I2C, use: sensor = VL53L5CXSensor(use_serial=False)
# Auto-detect ESP32 on serial port (quiet mode):
try:
    # Use the ToF sensor via ESP32 serial output.
    # Set verbose=False for clean output; enable for debugging if needed.
    sensor = VL53L5CXSensor(port=None, baudrate=250000, use_serial=True, verbose=False)
except Exception as e:
    # permission or port errors are expected if device busy; warn once
    print(f"ALERT: Could not initialize sensor: {e}")
    sensor = None

OBSTACLE_CLASSES = [
    "bicycle", "car", "motorcycle", "bus", "train", "truck", "boat",
    "traffic light", "fire hydrant", "stop sign", "parking meter",
    "person","knife","chair","laptop","cell phone","remote","keyboard","mouse"
      # stairs requires custom model later
]

neg_OBSTACLE_CLASSES = [ #filtering OUT rather than filtering IN,
    "person"
]

# Distance threshold for considering something "close" (in mm)
CLOSE_THRESHOLD_MM = 1000

# How often we remind the user about the same object
ALERT_REMINDER_SECONDS = 1.0  # reduced from 5.0 for faster re-alerting

# If an object stays in roughly the same place, we consider it "the same object".
TRACKED_OBJECT_MOVE_THRESHOLD_PX = 80  # pixels for center movement
TRACKED_OBJECT_MAX_AGE = 10.0  # seconds before forgetting a tracked object

# Tracking objects over time so we can re-alert only after a reminder interval
tracked_objects = []  # each entry: {label, center, last_alert, last_seen}

# Sensor cells that were below CLOSE_THRESHOLD_MM in the most recent frame.
# A reading is only treated as real if it appears in TWO consecutive frames,
# which eliminates single-frame sensor spikes without altering displayed values.
prev_close_cells: set = set()

# Sensor grid visualization parameters
GRID_COLOR = (50, 50, 50)   # Dark gray for grid lines (used in sensor grid)
SENSOR_GRID_SIZE = 600  # Size of sensor grid window (square)
SENSOR_CELL_SIZE = SENSOR_GRID_SIZE // 8  # Each cell is 1/8 of the grid
SENSOR_MAX_DISTANCE = 3500  # Sensor reliable range ceiling in mm (used for colour scaling only)
SENSOR_VOID_VALUE = 0       # Sensor outputs 0 for invalid / out-of-range readings
SENSOR_STALE_TIMEOUT = 0.5  # seconds before we treat the data as stale

# Shared state between the main loop and the sensor polling thread
last_sensor_print_time = 0.0
sensor_warning_printed = False
sensor_parse_warning_printed = False
sensor_data_lock = threading.Lock()
last_sensor_data = np.zeros((8, 8), dtype=np.int32)  # 0 = no data yet (matches sensor invalid value)
last_sensor_update_time = time.time()

# Thread stop signal (clean shutdown)
sensor_stop_event = threading.Event()


def _normalize_sensor_data(raw_sensor_data):
    """Normalize sensor data into an 8x8 numpy int32 array.

    Supports:
    - numpy arrays (1D length 64 or 2D 8x8)
    - list/tuple data (flat or nested)
    - dicts with keys like 'distances', 'data', or a single-value mapping

    Returns None when the input cannot be interpreted.
    """
    if raw_sensor_data is None:
        return None

    # Unwrap common dict wrappers
    if isinstance(raw_sensor_data, dict):
        for key in ("distances", "distance_mm", "ranging_data", "data", "grid", "frame"):
            if key in raw_sensor_data:
                return _normalize_sensor_data(raw_sensor_data[key])
        # If dict contains a single entry, try that value
        if len(raw_sensor_data) == 1:
            return _normalize_sensor_data(next(iter(raw_sensor_data.values())))
        return None

    # Convert to numpy array where possible
    try:
        arr = np.asarray(raw_sensor_data)
    except Exception:
        return None

    # If the data is already 8x8, we're good.
    if arr.shape == (8, 8):
        return arr.astype(np.int32, copy=False)

    # If it's 1D with 64 values, reshape to 8x8
    if arr.ndim == 1 and arr.size == 64:
        try:
            return arr.reshape((8, 8)).astype(np.int32, copy=False)
        except Exception:
            pass

    # If total entries == 64, try reshaping regardless of dims
    if arr.size == 64:
        try:
            return arr.reshape((8, 8)).astype(np.int32, copy=False)
        except Exception:
            pass

    # As a fallback, flatten and pad/truncate to 64 values.
    flat = arr.flatten()
    if flat.size < 64:
        # Pad with 0 (sensor's own invalid marker) — do NOT substitute a synthetic distance
        pad = np.zeros(64 - flat.size, dtype=np.int32)
        flat = np.concatenate([flat, pad])
    elif flat.size > 64:
        flat = flat[:64]

    # Ensure integer values; replace any non-finite floats with 0 (invalid marker)
    try:
        flat = flat.astype(np.int32, copy=False)
    except Exception:
        flat = np.array(flat, dtype=np.int32)

    flat = np.where(np.isfinite(flat.astype(float)), flat, 0).astype(np.int32)
    return flat.reshape((8, 8))


def _sensor_polling_thread():
    """Continuously poll the VL53L5CX sensor in a background thread.

    This decouples sensor I/O from the camera/YOLO loop so the camera stays smooth
    even if serial reads are slow or intermittent.
    Raw sensor values are written directly — no smoothing or value substitution.
    """
    global last_sensor_data, last_sensor_update_time, sensor_warning_printed, sensor_parse_warning_printed

    while not sensor_stop_event.is_set():
        if not sensor:
            # If sensor isn't initialized, avoid busy looping
            time.sleep(0.1)
            continue

        new_frame = sensor.get_ranging_data()
        if new_frame is None:
            # No new data: sleep briefly to avoid burning CPU
            time.sleep(0.005)
            continue

        last_sensor_update_time = time.time()

        normalized = _normalize_sensor_data(new_frame)
        if normalized is None:
            # Avoid spamming the console if we cannot parse sensor data
            if not sensor_parse_warning_printed:
                print("ALERT: sensor returned unexpected data format. Unable to normalize to 8x8 array.")
                sensor_parse_warning_printed = True
            continue

        # Write raw frame directly — no smoothing applied
        with sensor_data_lock:
            last_sensor_data = normalized

        # If the sensor is returning all-zero values, warn once (sensor may not be ready)
        if not sensor_warning_printed:
            if np.all(normalized == 0):
                print("ALERT: sensor returning all-zero values. Check wiring/position.")
                sensor_warning_printed = True


# Start background thread to keep sensor readings as fresh as possible
sensor_thread = threading.Thread(target=_sensor_polling_thread, daemon=True)
sensor_thread.start()


def create_sensor_grid(sensor_data):
    """
    Create a visualization grid for the 8x8 TOF sensor data
    Color-codes cells based on distance (closer = red/orange, farther = blue/green)
    """
    canvas = np.zeros((SENSOR_GRID_SIZE, SENSOR_GRID_SIZE, 3), dtype=np.uint8)
    
    # Draw grid lines
    for i in range(9):
        x = i * SENSOR_CELL_SIZE
        cv2.line(canvas, (x, 0), (x, SENSOR_GRID_SIZE), GRID_COLOR, 2)
        cv2.line(canvas, (0, x), (SENSOR_GRID_SIZE, x), GRID_COLOR, 2)
    
    # Fill cells with color based on distance
    for row in range(8):
        for col in range(8):
            distance = sensor_data[row, col]
            
            # Calculate cell position
            x1 = col * SENSOR_CELL_SIZE
            y1 = row * SENSOR_CELL_SIZE
            x2 = (col + 1) * SENSOR_CELL_SIZE
            y2 = (row + 1) * SENSOR_CELL_SIZE
            
            # Color mapping: closer objects = red/orange, farther = blue/green
            # Use raw distance in mm (0-3300), clamp to max distance
            clamped_dist = min(distance, SENSOR_MAX_DISTANCE)
            
            if distance == 0:
                # Invalid - gray
                color = (50, 50, 50)
            else:
                # Color gradient: red (close) -> yellow -> green -> blue (far)
                # Map 0-3300 to 0-1
                normalized_dist = clamped_dist / SENSOR_MAX_DISTANCE
                
                if normalized_dist < 0.25:
                    # Close: Red to Orange
                    r = 255
                    g = int(255 * (normalized_dist / 0.25))
                    b = 0
                elif normalized_dist < 0.5:
                    # Medium-close: Orange to Yellow
                    r = 255
                    g = 255
                    b = int(255 * ((normalized_dist - 0.25) / 0.25))
                elif normalized_dist < 0.75:
                    # Medium-far: Yellow to Green
                    r = int(255 * (1 - (normalized_dist - 0.5) / 0.25))
                    g = 255
                    b = 0
                else:
                    # Far: Green to Blue
                    r = 0
                    g = int(255 * (1 - (normalized_dist - 0.75) / 0.25))
                    b = int(255 * ((normalized_dist - 0.75) / 0.25))
                
                color = (int(b), int(g), int(r))  # BGR format for OpenCV
            
            # Fill cell with color
            cv2.rectangle(canvas, (x1 + 1, y1 + 1), (x2 - 1, y2 - 1), color, -1)
            
            # Add distance text (in mm, or '---' if sensor reports 0 / invalid)
            if distance == 0:
                text = "---"
            else:
                text = f"{int(distance)}"
            
            # Calculate text position (centered in cell)
            text_size = cv2.getTextSize(text, cv2.FONT_HERSHEY_SIMPLEX, 0.4, 1)[0]
            text_x = x1 + (SENSOR_CELL_SIZE - text_size[0]) // 2
            text_y = y1 + (SENSOR_CELL_SIZE + text_size[1]) // 2
            
            # Use white or black text based on cell brightness
            brightness = sum(color) / 3
            text_color = (255, 255, 255) if brightness < 128 else (0, 0, 0)
            
            cv2.putText(canvas, text, (text_x, text_y),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.4, text_color, 1)
    
    # Add title
    cv2.putText(canvas, "TOF Sensor Grid (mm)", (10, 25),
                cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)
    
    return canvas

def map_camera_to_sensor_grid(x1, y1, x2, y2, frame_height, frame_width):
    """
    Map camera frame bounding box coordinates to sensor grid coordinates (8x8)
    Assumes sensor is centered and aligned with camera view
    Returns list of (row, col) tuples for sensor cells that overlap with bbox
    """
    # Normalize bbox to 0-1 range
    norm_x1 = x1 / frame_width
    norm_y1 = y1 / frame_height
    norm_x2 = x2 / frame_width
    norm_y2 = y2 / frame_height
    
    # Map to 8x8 sensor grid (0-7 for rows and cols)
    sensor_col1 = int(norm_x1 * 8)
    sensor_row1 = int(norm_y1 * 8)
    sensor_col2 = int(norm_x2 * 8)
    sensor_row2 = int(norm_y2 * 8)
    
    # Clamp to valid range
    sensor_col1 = max(0, min(sensor_col1, 7))
    sensor_row1 = max(0, min(sensor_row1, 7))
    sensor_col2 = max(0, min(sensor_col2, 7))
    sensor_row2 = max(0, min(sensor_row2, 7))
    
    # Generate list of sensor cells covering the bbox
    sensor_cells = []
    for row in range(sensor_row1, sensor_row2 + 1):
        for col in range(sensor_col1, sensor_col2 + 1):
            sensor_cells.append((row, col))
    
    return sensor_cells

def check_sensor_close_in_region(sensor_data, sensor_cells, distance_threshold=400):
    """
    Check if sensor detected an object within distance_threshold (mm) in the specified region
    Returns tuple: (detected_close: bool, min_distance: int or None)
    """
    if sensor_data is None or len(sensor_cells) == 0:
        return False, None
    
    distances = []
    for row, col in sensor_cells:
        distance = sensor_data[row, col]
        # Valid distance reading
        if distance > 0:
            distances.append(distance)
    
    if len(distances) == 0:
        return False, None
    
    min_dist = min(distances)
    detected_close = min_dist < distance_threshold
    return detected_close, min_dist


# last_sensor_data is initialized and updated by the background sensor polling thread.

# (Debug printing disabled: only alerts and audio should be emitted)

# Excel logging setup (optional)
excel_filename = f"detection_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
wb = None
ws = None
ws_close = None  # second sheet: confirmed-close detections only
if _has_openpyxl:
    wb = Workbook()
    ws = wb.active
    ws.title = "Detections"

    # Create header row with styling (one row per detection)
    headers = [
        "Timestamp",
        "Object Class",
        "Confidence",
        "Location (px)",
        "Sensor Distance (mm)",
        "Sensor Cells",
        "Closest Cell",
        "Is Close",
        "Frame Avg Confidence",
        "Frame Close Count",
        "Frame Close Objects"
    ]
    ws.append(headers)

    # Style header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Column widths
    ws.column_dimensions['A'].width = 25  # Timestamp
    ws.column_dimensions['B'].width = 18  # Object class
    ws.column_dimensions['C'].width = 12  # Confidence
    ws.column_dimensions['D'].width = 18  # Location
    ws.column_dimensions['E'].width = 20  # Sensor distance
    ws.column_dimensions['F'].width = 30  # Sensor cells
    ws.column_dimensions['G'].width = 15  # Closest cell
    ws.column_dimensions['H'].width = 10  # Is close
    ws.column_dimensions['I'].width = 18  # Frame average conf
    ws.column_dimensions['J'].width = 15  # Frame close count
    ws.column_dimensions['K'].width = 40  # Frame close objects

    # ── Second sheet: one row per confirmed-close detection event ──────────────
    ws_close = wb.create_sheet(title="Close Detections")
    close_headers = [
        "Timestamp",
        "Object Class",
        "Sensor Distance (mm)",
        "YOLO Confidence",
    ]
    ws_close.append(close_headers)

    close_header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    for cell in ws_close[1]:
        cell.fill = close_header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws_close.column_dimensions['A'].width = 25  # Timestamp
    ws_close.column_dimensions['B'].width = 18  # Object class
    ws_close.column_dimensions['C'].width = 22  # Sensor distance
    ws_close.column_dimensions['D'].width = 18  # YOLO confidence

# Tracking for 0.25 second interval data logging
last_log_time = time.time()
LOG_INTERVAL = 0.25  # seconds
current_frame_detections = []

# optional auto-termination to prevent unresponsive pop-ups during testing
MAX_RUNTIME = 30  # seconds; set to None to run indefinitely
start_time = time.time()

while True:
    if cap is None:
        # No camera available; continue running using a black placeholder frame
        frame = np.zeros((480, 640, 3), dtype=np.uint8)
        ret = True
    else:
        ret, frame = cap.read()
        if not ret:
            print("ALERT: camera frame grab failed, exiting")
            break

    # Camera should be used with native orientation to match sensor mapping
    # (flipping caused sensor/camera mismatch)
    # frame = cv2.flip(frame, 1)

    # Copy latest sensor data (thread-safe) for visualization/detection
    with sensor_data_lock:
        sensor_data = last_sensor_data.copy()
        last_update = last_sensor_update_time

    # If the sensor has not updated in a while, treat it as stale (show all-zero / no-data)
    if time.time() - last_update > SENSOR_STALE_TIMEOUT:
        sensor_data = np.zeros((8, 8), dtype=np.int32)
    
    # Run YOLO inference on the frame
    results = model(frame, stream=True, verbose=False)
    
    # Track whether we've already spoken an alert this frame (one audio at a time)
    audio_alert_sent = False
    frame_now = time.time()
    _cleanup_tracked_objects(frame_now)

    # Reset detection list for this frame
    current_frame_detections = []
    # Track sensor cells covered by a YOLO detection (for unidentifiable-object check)
    covered_sensor_cells: set = set()
    # Collect every cell that is close this frame; used to update prev_close_cells
    current_close_cells: set = set()

    for r in results:
        boxes = r.boxes
        for b in boxes:
            #bounding box locations USE FOR SENSOR
            x1, y1, x2, y2 = b.xyxy[0]
            cls = int(b.cls[0])
            conf = float(b.conf[0])

            center_x = int((x1 + x2) / 2)
            center_y = int((y1 + y2) / 2)

            #draws box on main frame
            cv2.rectangle(frame, (int(x1), int(y1)), (int(x2), int(y2)), (0,255,0), 2)

            label = model.names[cls]
            #if label in OBSTACLE_CLASSES and conf > 0.4:
            if label in OBSTACLE_CLASSES:
                location_str = f"({center_x}, {center_y})"
                
                # Map the full bounding box to sensor grid cells, then keep only
                # rows 0–3 (top half of the 8×8 ToF sensor).  Any box that doesn't
                # overlap the top sensor half will get an empty list and no alert.
                sensor_cells = map_camera_to_sensor_grid(x1, y1, x2, y2, frame.shape[0], frame.shape[1])
                sensor_cells = [(r, c) for r, c in sensor_cells if r < 4]
                sensor_close, sensor_distance = check_sensor_close_in_region(
                    sensor_data, sensor_cells, distance_threshold=CLOSE_THRESHOLD_MM)

                # Track which cells are close this frame (for next-frame confirmation)
                for _r, _c in sensor_cells:
                    if 0 < sensor_data[_r, _c] < CLOSE_THRESHOLD_MM:
                        current_close_cells.add((_r, _c))

                # Require the close reading to have been present last frame too
                # (eliminates single-frame sensor spikes; raw grid display is unaffected)
                sensor_close_confirmed = sensor_close and any(
                    cell in prev_close_cells for cell in sensor_cells
                    if 0 < sensor_data[cell[0], cell[1]] < CLOSE_THRESHOLD_MM
                )

                # Mark these cells as covered by a known YOLO object
                covered_sensor_cells.update(sensor_cells)
                
                # determine closest cell inside this region (if sensor data available)
                closest_cell = None
                if sensor_data is not None and sensor_cells:
                    min_dist = float('inf')
                    for row, col in sensor_cells:
                        d = sensor_data[row, col]
                        if 0 < d < min_dist:
                            min_dist = d
                            closest_cell = (row, col)
                
                current_frame_detections.append({
                    'label': label,
                    'confidence': conf,
                    'location': location_str,
                    'sensor_close': sensor_close,
                    'sensor_close_confirmed': sensor_close_confirmed,
                    'sensor_distance': sensor_distance,
                    'sensor_cells': sensor_cells,
                    'closest_cell': closest_cell
                })
                
                cv2.putText(frame, f"{label} ({conf:.1f})", (int(x1), int(y1)-10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0,255,0), 2)
                # Only alert when a close reading is confirmed across two consecutive frames
                if sensor_close_confirmed:
                    obj = _find_tracked_object(label, (center_x, center_y))
                    should_alert = False

                    if obj is None:
                        obj = {
                            "label": label,
                            "center": (center_x, center_y),
                            "last_alert": frame_now,
                            "last_seen": frame_now,
                        }
                        tracked_objects.append(obj)
                        should_alert = True
                    else:
                        # Update location and last seen time
                        moved = _euclidean_dist((center_x, center_y), obj["center"])
                        obj["center"] = (center_x, center_y)
                        obj["last_seen"] = frame_now

                        # Remind only if object hasn't moved much and cooldown passed
                        if (frame_now - obj.get("last_alert", 0)) > ALERT_REMINDER_SECONDS and moved < TRACKED_OBJECT_MOVE_THRESHOLD_PX:
                            should_alert = True
                            obj["last_alert"] = frame_now

                    if should_alert:
                        direction = get_direction_descriptor(center_x, center_y, frame.shape[1], frame.shape[0])
                        alert_msg = f"ALERT: close object detected - {label} at {sensor_distance}mm, cell {closest_cell}"
                        print(alert_msg)
                        if not audio_alert_sent:
                            speak_text(f"Alert: close object detected in {direction} - {label}")
                            audio_alert_sent = True

    # ── Unidentifiable-object check ──────────────────────────────────────────────
    # If the ToF sensor sees something close in the upper half (rows 0-3) but YOLO
    # did not identify any object there, alert as "unidentifiable object".
    # Confirmation required: cell must also have been close in the previous frame.
    unidentifiable_close = False
    min_unidentified_dist = None
    for row in range(4):          # upper half of the 8×8 sensor grid
        for col in range(8):
            if (row, col) in covered_sensor_cells:
                continue          # already accounted for by a known detection
            dist = int(sensor_data[row, col])
            if 0 < dist < (CLOSE_THRESHOLD_MM-200):
                current_close_cells.add((row, col))
                if (row, col) in prev_close_cells:   # confirmed across two frames
                    unidentifiable_close = True
                    if min_unidentified_dist is None or dist < min_unidentified_dist:
                        min_unidentified_dist = dist
    #UNIDENTIFIABLE OBJECT CHECK DISABLED
    #if unidentifiable_close:
        #print(f"ALERT: unidentifiable object detected at ~{min_unidentified_dist}mm (sensor, no YOLO match)")
        #if not audio_alert_sent:
            #speak_text("Alert: unidentifiable object detected")
            #audio_alert_sent = True

    # Advance the close-cell history for the next frame's confirmation check
    prev_close_cells = current_close_cells

    
    # Create sensor visualization grid (use last data, or empty grid if no data yet)
    if sensor_data is not None:
        sensor_grid = create_sensor_grid(sensor_data)
    else:
        # Create empty grid until first data arrives
        sensor_grid = np.zeros((SENSOR_GRID_SIZE, SENSOR_GRID_SIZE, 3), dtype=np.uint8)
        cv2.putText(sensor_grid, "Waiting for sensor data...", (SENSOR_GRID_SIZE//2 - 150, SENSOR_GRID_SIZE//2),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, (128, 128, 128), 2)
    
    cv2.imshow("Obstacle Detection Demo", frame)
    cv2.imshow("TOF Sensor Grid", sensor_grid)
    
    # Log data to Excel every 0.25 seconds (if available)
    if _has_openpyxl:
        current_time = time.time()
        if current_time - last_log_time >= LOG_INTERVAL:
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
            total_instances = len(current_frame_detections)
            
            # frame-level metrics
            if total_instances > 0:
                frame_avg_conf = sum(d['confidence'] for d in current_frame_detections) / total_instances
                close_objs = [d['label'] for d in current_frame_detections if d['sensor_close']]
                frame_close_count = len(close_objs)
                frame_close_list = ", ".join(close_objs)
            else:
                frame_avg_conf = 0
                frame_close_count = 0
                frame_close_list = ""
            
            # add one row per detection (main sheet)
            for d in current_frame_detections:
                cells_str = ", ".join([f"({r},{c})" for r, c in d['sensor_cells']]) if d['sensor_cells'] else ""
                is_close_str = "Y" if d['sensor_close'] else ""
                closest_cell_str = str(d['closest_cell']) if d.get('closest_cell') is not None else ""
                ws.append([
                    timestamp,
                    d['label'],
                    f"{d['confidence']:.2f}",
                    d['location'],
                    d['sensor_distance'] if d['sensor_distance'] is not None else "",
                    cells_str,
                    closest_cell_str,
                    is_close_str,
                    f"{frame_avg_conf:.2f}",
                    frame_close_count,
                    frame_close_list
                ])

            # "Close Detections" sheet — one row per confirmed-close event only
            if ws_close is not None:
                for d in current_frame_detections:
                    if d.get('sensor_close_confirmed') and d['sensor_distance'] is not None:
                        ws_close.append([
                            timestamp,
                            d['label'],
                            d['sensor_distance'],
                            round(d['confidence'], 2),
                        ])

            wb.save(excel_filename)
            last_log_time = current_time
    
    # window event handling and exit conditions
    if cv2.waitKey(1) == 27:
        break

# Cleanup
sensor_stop_event.set()
if sensor_thread.is_alive():
    sensor_thread.join(timeout=1.0)

cap.release()
cv2.destroyAllWindows()
if sensor:
    sensor.close()

# Save and close workbook (if logging enabled)
if _has_openpyxl and wb is not None:
    wb.save(excel_filename)




#if label in OBSTACLE_CLASSES and conf > 0.5:
    